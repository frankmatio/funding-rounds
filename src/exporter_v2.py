"""
Data Export Utilities V2
Export funding rounds from database to Excel, CSV, and JSON
"""

import csv
import json
import logging
import os
from datetime import datetime
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .database.models import FundingRound

logger = logging.getLogger(__name__)


class ExporterV2:
    """Export funding rounds from database to various formats"""

    def __init__(self, config: dict, db_manager=None):
        """Initialize exporter"""
        self.config = config
        self.db_manager = db_manager

        export_config = config.get('export', {})
        self.output_directory = export_config.get('output_directory', 'data/exports')
        self.filename_template = export_config.get('filename_template', 'funding_rounds_{timestamp}')

        # Create output directory
        os.makedirs(self.output_directory, exist_ok=True)

        logger.info(f"✓ Exporter initialized")
        logger.info(f"  Output directory: {self.output_directory}")

    def get_filename(self, extension: str) -> str:
        """Generate filename with timestamp"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = self.filename_template.format(timestamp=timestamp)
        return os.path.join(self.output_directory, f"{filename}.{extension}")

    def prepare_export_data(self, session) -> List[dict]:
        """
        Prepare funding rounds for export.
        Returns list of dicts with flattened data.
        """
        # Get all non-duplicate rounds
        rounds = session.query(FundingRound).filter_by(is_duplicate=False).all()

        export_data = []

        for round_obj in rounds:
            # Flatten investor data
            investor_names = [inv.name for inv in round_obj.investors] if round_obj.investors else []
            all_investors_str = ', '.join(investor_names) if investor_names else ''

            # Get source URLs
            source_urls = round_obj.source_urls if round_obj.source_urls else []
            source_urls_str = ', '.join(source_urls) if isinstance(source_urls, list) else str(source_urls)

            row = {
                'Company Name': round_obj.company.name,
                'Company CIK': round_obj.company.cik or '',
                'Round Name': round_obj.round_name or '',
                'Date': round_obj.date or '',
                'Amount Raised (USD)': round_obj.amount_raised_usd,
                'Pre-Money Valuation (USD)': round_obj.pre_money_valuation_usd,
                'Post-Money Valuation (USD)': round_obj.post_money_valuation_usd,
                'Lead Investor': round_obj.lead_investor or '',
                'All Investors': all_investors_str,
                'Source Type': round_obj.source_type,
                'Confidence Score': round_obj.confidence_score,
                'Source URLs': source_urls_str,
                'Notes': round_obj.notes or '',
            }

            export_data.append(row)

        return export_data

    def export_to_excel(self, session) -> str:
        """Export to Excel with formatting"""
        logger.info("Exporting to Excel...")

        data = self.prepare_export_data(session)

        if not data:
            logger.warning("No data to export")
            return None

        # Create DataFrame
        df = pd.DataFrame(data)

        # Generate filename
        filename = self.get_filename('xlsx')

        # Write to Excel
        df.to_excel(filename, index=False, sheet_name='Funding Rounds')

        # Apply formatting
        wb = load_workbook(filename)
        ws = wb.active

        # Header formatting
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save
        wb.save(filename)

        logger.info(f"✓ Excel export complete: {filename}")
        logger.info(f"  Rows: {len(data)}")

        return filename

    def export_to_csv(self, session) -> str:
        """Export to CSV"""
        logger.info("Exporting to CSV...")

        data = self.prepare_export_data(session)

        if not data:
            logger.warning("No data to export")
            return None

        # Generate filename
        filename = self.get_filename('csv')

        # Write to CSV
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            if data:
                writer = csv.DictWriter(f, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)

        logger.info(f"✓ CSV export complete: {filename}")
        logger.info(f"  Rows: {len(data)}")

        return filename

    def export_to_json(self, session) -> str:
        """Export to JSON"""
        logger.info("Exporting to JSON...")

        data = self.prepare_export_data(session)

        if not data:
            logger.warning("No data to export")
            return None

        # Generate filename
        filename = self.get_filename('json')

        # Write to JSON
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        logger.info(f"✓ JSON export complete: {filename}")
        logger.info(f"  Rows: {len(data)}")

        return filename

    def export_all_formats(self, session) -> dict:
        """Export to all configured formats"""
        logger.info("=" * 80)
        logger.info("EXPORTING DATA")
        logger.info("=" * 80)

        formats = self.config.get('export', {}).get('formats', ['excel', 'csv', 'json'])

        results = {}

        if 'excel' in formats:
            results['excel'] = self.export_to_excel(session)

        if 'csv' in formats:
            results['csv'] = self.export_to_csv(session)

        if 'json' in formats:
            results['json'] = self.export_to_json(session)

        logger.info("")
        logger.info("Export complete!")
        logger.info(f"Output directory: {self.output_directory}")
        logger.info("")

        return results
